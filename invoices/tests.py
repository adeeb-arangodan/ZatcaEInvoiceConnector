import json
import uuid
import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import ec as ec_module
from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from organization.models import Device, DeviceKeyMaterial, Organization
from organization.services import encrypt_private_key

User = get_user_model()

from .hashing import INITIAL_PIH, get_icv_and_pih_atomically, store_invoice_hash
from .models import InvoiceSubmission, InvoiceSubmissionFailure
from .pipeline import InvoiceSubmissionRejected, process_invoice_submission
from .serializers import InvoiceSubmissionSerializer
from .services import (
    DuplicateReturnNumberError,
    create_custom_return_credit_note,
    create_return_credit_note,
)
from .submission import submit_to_zatca
from .xml_builder import _compute_totals, build_compliance_sample_invoice

SUBMIT_URL = '/api/invoices/submit/'

VALID_PAYLOAD = {
    'device_asset_id': 'ASSET-100',
    'invoice_number': 'INV-001',
    'issue_date': '2026-06-09',
    'issue_time': '10:00:00',
    'invoice_type_code': '388',
    'invoice_type_code_name_attribute': '020000000',
    'customer_name': 'Test Customer',
    'customer_vat': '300000000000003',
    'customer_city': 'Riyadh',
    'customer_country_code': 'SA',
    'items': [
        {
            'slno': 1,
            'code': 'ITEM-001',
            'name': 'Consultation',
            'qty': '1.0000',
            'price': '100.0000',
            'vat_type': 'S',
        }
    ],
}

ORG_DEFAULTS = dict(
    name='Test Org',
    branch_name='Branch-1',
    industry_category='Healthcare',
    vat_number='399999999900003',
    country_code='SA',
    national_address_code='RCFA3435',
    street_name='Main St',
    building_number='1',
    city_sub_division='District',
    city_name='Riyadh',
    postal_zone='12345',
    cr_number='1010138184',
    invoice_category='1100',
    is_active=True,
)

FAKE_CSID = {
    # ZATCA's binarySecurityToken is base64-of-DER, base64-encoded again for
    # transport. This is a real (self-signed, secp256k1) test certificate so
    # signing.py's x509 parsing succeeds the same way it would for a real CSID.
    'binarySecurityToken': (
        'TUlJQkJqQ0JycUFEQWdFQ0FnRUJNQW9HQ0NxR1NNNDlCQU1DTUE4eERUQUxCZ05WQkFNTUJGUkZVMVF3'
        'SGhjTk1qUXdNVEF4TURBd01EQXdXaGNOTXpBd01UQXhNREF3TURBd1dqQVBNUTB3Q3dZRFZRUUREQVJV'
        'UlZOVU1GWXdFQVlIS29aSXpqMENBUVlGSzRFRUFBb0RRZ0FFNEpZSUluT1BaQWQ3eDlKZnFHZVVnVjNN'
        'Y2VDcTVQVW1HNndiL2Q0MkQ0MzZxSlRoRWMvQStZVFk5Z3E3OTJJWWI4QVczcWw3dkVuWllmaUZJVzFt'
        'N2pBS0JnZ3Foa2pPUFFRREFnTkhBREJFQWlCYzI4eWJDK3JNWjlMV3RZZ01KUjBENk9yd3pTU2V4ZzlT'
        'TnhPWEpOakN6QUlnVm1qZGk3MWxTYzdtV25CZHllZ0dzQTJWZW1ENWxRS0xFQkNaeStKdi81cz0='
    ),
    'secret': 'testsecret',
    'requestID': 'REQ-001',
}


def _make_stub_submission(**kwargs):
    stub = MagicMock(spec=InvoiceSubmission)
    stub.pk = 1
    stub.status = 'submitted'
    stub.qr_code_data = 'AQID'
    stub.invoice_uuid = uuid.uuid4()
    stub.zatca_response = {'status_code': 200}
    for k, v in kwargs.items():
        setattr(stub, k, v)
    return stub


class InvoiceSubmitViewTests(TestCase):

    def _make_org_with_device(self, **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        org = Organization.objects.create(**defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        return org, device

    def _auth_header(self, org):
        return {'HTTP_AUTHORIZATION': f'ApiKey {org.api_key}'}

    def _post(self, payload, org=None, extra_headers=None):
        headers = self._auth_header(org) if org else {}
        if extra_headers:
            headers.update(extra_headers)
        return self.client.post(
            SUBMIT_URL,
            data=json.dumps(payload),
            content_type='application/json',
            **headers,
        )

    @patch('invoices.views.process_invoice_submission')
    def test_valid_invoice_returns_201(self, mock_pipeline):
        org, _ = self._make_org_with_device()
        mock_pipeline.return_value = _make_stub_submission()
        response = self._post(VALID_PAYLOAD, org)
        self.assertEqual(response.status_code, 201)
        body = response.json()
        self.assertIn('id', body)
        self.assertIn('qr_code', body)
        self.assertIn('invoice_uuid', body)
        self.assertEqual(body['status'], 'submitted')

    def test_missing_api_key_returns_401(self):
        response = self.client.post(
            SUBMIT_URL,
            data=json.dumps(VALID_PAYLOAD),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 401)

    def test_invalid_api_key_returns_401(self):
        response = self.client.post(
            SUBMIT_URL,
            data=json.dumps(VALID_PAYLOAD),
            content_type='application/json',
            HTTP_AUTHORIZATION='ApiKey totally-invalid-key',
        )
        self.assertEqual(response.status_code, 401)

    def test_inactive_org_returns_403(self):
        org, _ = self._make_org_with_device(is_active=False)
        response = self._post(VALID_PAYLOAD, org)
        self.assertEqual(response.status_code, 403)

    def test_missing_invoice_number_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {**VALID_PAYLOAD}
        del payload['invoice_number']
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('invoice_number', response.json())

    def test_device_from_other_org_returns_400(self):
        self._make_org_with_device()
        other_org = Organization.objects.create(
            name='Other Org', branch_name='B', industry_category='IT',
            vat_number='399999999900004', country_code='SA',
            national_address_code='X', street_name='Y', building_number='2',
            city_sub_division='D', city_name='Jeddah', postal_zone='11111',
            cr_number='9999999999', invoice_category='1100', is_active=True,
        )
        response = self._post(VALID_PAYLOAD, other_org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('device_asset_id', response.json())

    def test_credit_note_without_billing_reference_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {**VALID_PAYLOAD, 'invoice_type_code': '381', 'invoice_type_code_name_attribute': '020000000'}
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('billing_reference', response.json())

    def test_debit_note_without_billing_reference_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {**VALID_PAYLOAD, 'invoice_type_code': '383', 'invoice_type_code_name_attribute': '020000000'}
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('billing_reference', response.json())

    def test_credit_note_without_reason_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {
            **VALID_PAYLOAD,
            'invoice_type_code': '381',
            'invoice_type_code_name_attribute': '020000000',
            'billing_reference': 'INV-001',
        }
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('reason', response.json())

    def test_empty_items_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {**VALID_PAYLOAD, 'items': []}
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('items', response.json())

    def test_hea_exemption_without_customer_id_number_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {
            **VALID_PAYLOAD,
            'items': [{
                **VALID_PAYLOAD['items'][0],
                'vat_type': 'Z',
                'VatExceptionReason': 'VATEX-SA-HEA',
            }],
        }
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('customer_id_number', response.json())

    def test_edu_exemption_without_customer_id_number_returns_400(self):
        org, _ = self._make_org_with_device()
        payload = {
            **VALID_PAYLOAD,
            'items': [{
                **VALID_PAYLOAD['items'][0],
                'vat_type': 'Z',
                'VatExceptionReason': 'VATEX-SA-EDU',
            }],
        }
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 400)
        self.assertIn('customer_id_number', response.json())

    @patch('invoices.views.process_invoice_submission')
    def test_hea_exemption_with_customer_id_number_returns_201(self, mock_pipeline):
        org, _ = self._make_org_with_device()
        mock_pipeline.return_value = _make_stub_submission()
        payload = {
            **VALID_PAYLOAD,
            'customer_id_number': '1234567890',
            'items': [{
                **VALID_PAYLOAD['items'][0],
                'vat_type': 'Z',
                'VatExceptionReason': 'VATEX-SA-HEA',
            }],
        }
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 201)

    def test_duplicate_invoice_number_for_same_org_and_type_returns_400(self):
        org, _ = self._make_org_with_device()
        InvoiceSubmission.objects.create(
            organization=org,
            device=Device.objects.get(organization=org),
            document_type=InvoiceSubmission.DOCUMENT_TYPE_INVOICE,
            invoice_number=VALID_PAYLOAD['invoice_number'],
            payload={},
            icv=1,
        )

        response = self._post(VALID_PAYLOAD, org)

        self.assertEqual(response.status_code, 400)
        self.assertIn('invoice_number', response.json())

    def test_same_invoice_number_allowed_for_different_document_types(self):
        org, _ = self._make_org_with_device()
        InvoiceSubmission.objects.create(
            organization=org,
            device=Device.objects.get(organization=org),
            document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE,
            invoice_number=VALID_PAYLOAD['invoice_number'],
            payload={},
            icv=1,
        )

        with patch('invoices.views.process_invoice_submission') as mock_pipeline:
            mock_pipeline.return_value = _make_stub_submission()
            response = self._post(VALID_PAYLOAD, org)

        self.assertEqual(response.status_code, 201)

    def test_same_invoice_number_allowed_for_different_organizations(self):
        org, _ = self._make_org_with_device()
        other_org = Organization.objects.create(
            name='Other Org', branch_name='B', industry_category='IT',
            vat_number='399999999900066', country_code='SA',
            national_address_code='X', street_name='Y', building_number='2',
            city_sub_division='D', city_name='Jeddah', postal_zone='11111',
            cr_number='9999999996', invoice_category='1100', is_active=True,
        )
        InvoiceSubmission.objects.create(
            organization=other_org,
            device=Device.objects.create(
                organization=other_org, asset_id='OTHER-DEV', egs_sw_serial_number='S', otp='1',
            ),
            document_type=InvoiceSubmission.DOCUMENT_TYPE_INVOICE,
            invoice_number=VALID_PAYLOAD['invoice_number'],
            payload={},
            icv=1,
        )

        with patch('invoices.views.process_invoice_submission') as mock_pipeline:
            mock_pipeline.return_value = _make_stub_submission()
            response = self._post(VALID_PAYLOAD, org)

        self.assertEqual(response.status_code, 201)

    @patch('invoices.views.process_invoice_submission')
    def test_submission_calls_pipeline_with_correct_args(self, mock_pipeline):
        org, device = self._make_org_with_device()
        mock_pipeline.return_value = _make_stub_submission()
        self._post(VALID_PAYLOAD, org)
        self.assertTrue(mock_pipeline.called)
        call_kwargs = mock_pipeline.call_args
        self.assertEqual(call_kwargs.kwargs['organization'], org)
        self.assertEqual(call_kwargs.kwargs['device'], device)

    @patch('invoices.views.process_invoice_submission')
    def test_credit_note_with_billing_reference_returns_201(self, mock_pipeline):
        org, _ = self._make_org_with_device()
        mock_pipeline.return_value = _make_stub_submission()
        payload = {
            **VALID_PAYLOAD,
            'invoice_type_code': '381',
            'invoice_type_code_name_attribute': '020000000',
            'billing_reference': 'INV-001',
            'reason': 'Goods returned',
        }
        response = self._post(payload, org)
        self.assertEqual(response.status_code, 201)

    def test_device_without_csid_returns_422(self):
        org, device = self._make_org_with_device()
        device.csid_response = None
        device.save()
        response = self._post(VALID_PAYLOAD, org)
        self.assertEqual(response.status_code, 422)


class OrganizationApiKeyTests(TestCase):

    def _make_org(self, **overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(overrides)
        return Organization.objects.create(**defaults)

    def test_api_key_auto_generated_on_create(self):
        org = self._make_org()
        self.assertTrue(org.api_key)
        self.assertEqual(len(org.api_key), 64)

    def test_api_key_not_regenerated_on_update(self):
        org = self._make_org()
        original_key = org.api_key
        org.city_name = 'Jeddah'
        org.save()
        org.refresh_from_db()
        self.assertEqual(org.api_key, original_key)

    def test_api_key_is_unique_across_orgs(self):
        org1 = self._make_org()
        org2 = self._make_org(vat_number='399999999900004', cr_number='9999999999')
        self.assertNotEqual(org1.api_key, org2.api_key)


class InvoiceHashingTests(TestCase):

    def _make_org(self, **overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(overrides)
        return Organization.objects.create(**defaults)

    def test_first_call_returns_icv_one_and_initial_pih(self):
        org = self._make_org()

        icv, pih = get_icv_and_pih_atomically(org)

        self.assertEqual(icv, 1)
        self.assertEqual(pih, INITIAL_PIH)

    def test_second_call_returns_icv_two_and_stored_hash(self):
        org = self._make_org()

        get_icv_and_pih_atomically(org)
        store_invoice_hash(org, 'first-invoice-hash')
        icv, pih = get_icv_and_pih_atomically(org)

        self.assertEqual(icv, 2)
        self.assertEqual(pih, 'first-invoice-hash')

    def test_counter_is_shared_across_devices_in_same_organization(self):
        org = self._make_org()
        device_a = Device.objects.create(
            organization=org,
            asset_id='ASSET-A',
            egs_sw_serial_number='SERIAL-A',
            otp='111111',
        )
        device_b = Device.objects.create(
            organization=org,
            asset_id='ASSET-B',
            egs_sw_serial_number='SERIAL-B',
            otp='222222',
        )

        icv_a, pih_a = get_icv_and_pih_atomically(device_a.organization)
        store_invoice_hash(device_a.organization, 'hash-from-device-a')
        icv_b, pih_b = get_icv_and_pih_atomically(device_b.organization)

        self.assertEqual(icv_a, 1)
        self.assertEqual(pih_a, INITIAL_PIH)
        self.assertEqual(icv_b, 2)
        self.assertEqual(pih_b, 'hash-from-device-a')

    def test_store_invoice_hash_persists_on_organization(self):
        org = self._make_org()

        store_invoice_hash(org, 'some-hash')
        org.refresh_from_db()

        self.assertEqual(org.last_invoice_hash, 'some-hash')


class ComputeTotalsRoundingTests(TestCase):
    """Regression coverage for BR-CO-15: TaxInclusiveAmount (BT-112) must equal
    the already-rounded TaxExclusiveAmount (BT-109) + TaxAmount (BT-110). Rounding
    their unrounded sum independently can drift by a cent — reproduced concretely
    below and confirmed to affect ~24.6% of possible line-extension amounts before
    the fix."""

    def _totals(self, price, qty='1.0000', vat_type='S', advance_paid=0):
        items = [{'slno': 1, 'code': 'ITEM-001', 'name': 'Item', 'qty': qty, 'price': price, 'vat_type': vat_type}]
        return _compute_totals(items, advance_paid=advance_paid)

    def test_boundary_case_satisfies_br_co_15(self):
        totals = self._totals(price='0.0044')

        self.assertEqual(totals['tax_exclusive'], Decimal('0.00'))
        self.assertEqual(totals['vat_total'], Decimal('0.00'))
        self.assertEqual(totals['tax_inclusive'], totals['tax_exclusive'] + totals['vat_total'])

    def test_br_co_15_holds_across_a_sweep_of_prices(self):
        for cents in range(0, 20000, 7):
            price = str(Decimal(cents) / Decimal(10000))
            totals = self._totals(price=price)
            self.assertEqual(
                totals['tax_inclusive'], totals['tax_exclusive'] + totals['vat_total'],
                f'BR-CO-15 violated for price={price}',
            )

    def test_payable_derives_from_rounded_tax_inclusive_and_advance(self):
        totals = self._totals(price='0.0044', advance_paid='0.005')

        self.assertEqual(totals['payable'], totals['tax_inclusive'] - totals['advance'])

    def test_whole_number_amounts_unaffected(self):
        totals = self._totals(price='100.0000')

        self.assertEqual(totals['tax_exclusive'], Decimal('100.00'))
        self.assertEqual(totals['vat_total'], Decimal('15.00'))
        self.assertEqual(totals['tax_inclusive'], Decimal('115.00'))


@override_settings(DEVICE_KEY_ENCRYPTION_KEY=Fernet.generate_key().decode())
class InvoicePipelineTests(TestCase):
    """Exercises process_invoice_submission() end-to-end (real XML build/hash/sign), mocking only submit_to_zatca."""

    def _make_org_with_signing_device(self, **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        org = Organization.objects.create(**defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device

    def _validated(self, payload, org):
        serializer = InvoiceSubmissionSerializer(data=payload, organization=org)
        serializer.is_valid(raise_exception=True)
        return serializer.validated_data, serializer.get_resolved_device()

    @patch('invoices.pipeline.submit_to_zatca')
    def test_chain_does_not_advance_when_zatca_rejects(self, mock_submit):
        """ZATCA tracks the last ICV/hash it accepted on its own side — advancing our
        chain on an invoice it rejects would leave a gap it can never reconcile, and
        every later invoice would inherit that gap and get rejected too. So a rejection
        must roll back the whole attempt: no row, no ICV/PIH advance."""
        org, device = self._make_org_with_signing_device()
        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)

        with self.assertRaises(InvoiceSubmissionRejected) as ctx:
            process_invoice_submission(org, resolved_device, validated_data)

        org.refresh_from_db()
        self.assertEqual(org.invoice_counter, 0)
        self.assertEqual(org.last_invoice_hash, '')
        self.assertEqual(InvoiceSubmission.objects.count(), 0)

        failure = ctx.exception.failure
        self.assertEqual(InvoiceSubmissionFailure.objects.count(), 1)
        self.assertEqual(failure.organization, org)
        self.assertEqual(failure.device, device)
        self.assertEqual(failure.document_type, InvoiceSubmission.DOCUMENT_TYPE_INVOICE)
        self.assertEqual(failure.invoice_number, VALID_PAYLOAD['invoice_number'])
        self.assertEqual(failure.zatca_response, {'status_code': 422, 'error': {'message': 'invalid'}})
        self.assertFalse(failure.resolved)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_retry_after_rejection_reuses_same_icv(self, mock_submit):
        org, device = self._make_org_with_signing_device()
        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)

        with self.assertRaises(InvoiceSubmissionRejected):
            process_invoice_submission(org, resolved_device, validated_data)

        mock_submit.return_value = {'status_code': 200}
        submission = process_invoice_submission(org, resolved_device, validated_data)

        self.assertEqual(submission.icv, 1)
        self.assertEqual(submission.status, InvoiceSubmission.STATUS_SUBMITTED)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_chain_advances_on_zatca_acceptance(self, mock_submit):
        org, device = self._make_org_with_signing_device()
        mock_submit.return_value = {'status_code': 200}
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)

        submission = process_invoice_submission(org, resolved_device, validated_data)

        self.assertEqual(submission.status, InvoiceSubmission.STATUS_SUBMITTED)
        self.assertIsNotNone(submission.submitted_at)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_pih_is_not_committed_until_zatca_accepts(self, mock_submit):
        """Regression test for the inverted invariant: the org row lock (and
        last_invoice_hash) must NOT be released/stored until ZATCA has actually
        accepted the invoice — otherwise a later rejection would leave a chain gap
        ZATCA can never reconcile. last_invoice_hash must still hold the pre-submission
        value at the moment ZATCA is contacted."""
        org, device = self._make_org_with_signing_device()
        seen = {}

        def fake_submit(*args, **kwargs):
            seen['last_invoice_hash'] = Organization.objects.get(pk=org.pk).last_invoice_hash
            return {'status_code': 200}

        mock_submit.side_effect = fake_submit
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)

        submission = process_invoice_submission(org, resolved_device, validated_data)

        self.assertEqual(seen['last_invoice_hash'], '')
        self.assertNotEqual(submission.invoice_hash, seen['last_invoice_hash'])

    @patch('invoices.pipeline.submit_to_zatca')
    def test_build_failure_rolls_back_icv_and_leaves_no_row(self, mock_submit):
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)

        with patch('invoices.pipeline.build_invoice_xml', side_effect=ValueError('boom')):
            with self.assertRaises(ValueError):
                process_invoice_submission(org, resolved_device, validated_data)

        org.refresh_from_db()
        self.assertEqual(org.invoice_counter, 0)
        self.assertEqual(org.last_invoice_hash, '')
        self.assertEqual(InvoiceSubmission.objects.count(), 0)
        mock_submit.assert_not_called()


class ComplianceSampleInvoiceTests(TestCase):
    def _make_device(self):
        org = Organization.objects.create(**ORG_DEFAULTS)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        return device

    def test_build_compliance_sample_invoice_uses_given_type_and_billing_fields(self):
        device = self._make_device()

        xml_bytes, invoice_uuid, invoice_hash, fake_data = build_compliance_sample_invoice(
            device,
            invoice_type_code='381',
            name_attribute='010000000',
            billing_reference='REF-1',
            reason='why',
        )

        xml_text = xml_bytes.decode('utf-8')
        self.assertIn('<cbc:InvoiceTypeCode name="010000000">381</cbc:InvoiceTypeCode>', xml_text)
        self.assertIn('<cbc:ID>REF-1</cbc:ID>', xml_text)
        self.assertIn('<cbc:InstructionNote>why</cbc:InstructionNote>', xml_text)
        self.assertIsNotNone(invoice_uuid)
        self.assertTrue(invoice_hash)
        self.assertEqual(fake_data['invoice_type_code'], '381')
        self.assertEqual(fake_data['billing_reference'], 'REF-1')

    def test_build_compliance_sample_invoice_defaults_to_simplified_invoice(self):
        device = self._make_device()

        xml_bytes, _, _, _ = build_compliance_sample_invoice(device)

        xml_text = xml_bytes.decode('utf-8')
        self.assertIn('<cbc:InvoiceTypeCode name="020000000">388</cbc:InvoiceTypeCode>', xml_text)
        self.assertNotIn('<cac:BillingReference>', xml_text)


class SubmitToZatcaRoutingTests(TestCase):
    """Confirms KSA-2 routing: '02...' -> reporting (simplified), '01...' -> clearance
    (standard). Both subtypes start with '0', so this guards against regressing to a
    startswith('0') check that would route every invoice to reporting."""

    def _make_device(self):
        org = Organization.objects.create(**ORG_DEFAULTS)
        return Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )

    def _post_and_capture_url(self, device, name_attribute):
        captured = {}

        class DummyRequests:
            class HTTPError(Exception):
                pass

            class RequestException(Exception):
                pass

            @staticmethod
            def post(url, headers, json, timeout):
                captured['url'] = url

                class DummyResponse:
                    status_code = 200

                    @staticmethod
                    def raise_for_status():
                        return None

                    @staticmethod
                    def json():
                        return {}

                return DummyResponse()

        with patch('invoices.submission._get_requests_module', return_value=DummyRequests):
            submit_to_zatca(device, 'HASH', str(uuid.uuid4()), 'ENCODED', name_attribute)

        return captured['url']

    def test_simplified_code_routes_to_reporting_endpoint(self):
        device = self._make_device()

        url = self._post_and_capture_url(device, '020000000')

        self.assertTrue(url.endswith(settings.ZATCA_REPORTING_API_ENDPOINT))

    def test_standard_code_routes_to_clearance_endpoint(self):
        device = self._make_device()

        url = self._post_and_capture_url(device, '010000000')

        self.assertTrue(url.endswith(settings.ZATCA_CLEARANCE_API_ENDPOINT))


class SubmissionRejectionCallSiteTests(TestCase):
    """Exercises the real (unmocked) pipeline rejection path through each of the
    three HTTP-facing call sites, confirming they handle InvoiceSubmissionRejected
    instead of assuming a returned submission."""

    def _make_org_with_signing_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device, user

    def _auth_header(self, org):
        return {'HTTP_AUTHORIZATION': f'ApiKey {org.api_key}'}

    @patch('invoices.pipeline.submit_to_zatca')
    def test_invoice_submit_view_rejection_returns_422_and_creates_no_row(self, mock_submit):
        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        org, device, _user = self._make_org_with_signing_device()

        response = self.client.post(
            SUBMIT_URL, data=json.dumps(VALID_PAYLOAD), content_type='application/json',
            **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn('failure_id', response.json())
        self.assertEqual(InvoiceSubmission.objects.count(), 0)
        self.assertEqual(InvoiceSubmissionFailure.objects.count(), 1)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_invoice_return_view_rejection_returns_422_and_creates_no_row(self, mock_submit):
        org, device, _user = self._make_org_with_signing_device()
        mock_submit.return_value = {'status_code': 200}
        serializer = InvoiceSubmissionSerializer(data=VALID_PAYLOAD, organization=org)
        serializer.is_valid(raise_exception=True)
        invoice = process_invoice_submission(org, serializer.get_resolved_device(), serializer.validated_data)

        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        response = self.client.post(
            f'/api/invoices/{invoice.pk}/return/',
            data=json.dumps({'reason': 'test'}),
            content_type='application/json',
            **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn('failure_id', response.json())
        self.assertEqual(
            InvoiceSubmission.objects.filter(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE).count(), 0,
        )

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_invoice_form_view_rejection_shows_message(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        mock_submit.return_value = {'status_code': 200}
        serializer = InvoiceSubmissionSerializer(data=VALID_PAYLOAD, organization=org)
        serializer.is_valid(raise_exception=True)
        invoice = process_invoice_submission(org, serializer.get_resolved_device(), serializer.validated_data)

        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        self.client.force_login(user)
        response = self.client.post(
            reverse('organization:invoice-return', args=[org.pk, invoice.pk]),
            {'reason': 'damaged'},
        )

        self.assertRedirects(response, reverse('organization:invoice-list', args=[org.pk]))
        self.assertEqual(
            InvoiceSubmission.objects.filter(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE).count(), 0,
        )


@override_settings(DEVICE_KEY_ENCRYPTION_KEY=Fernet.generate_key().decode())
class DocumentRoutingTests(TestCase):
    """Confirms invoice_type_code sets the correct document_type on the shared table."""

    def _make_org_with_signing_device(self, **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        org = Organization.objects.create(**defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device

    def _validated(self, payload, org):
        serializer = InvoiceSubmissionSerializer(data=payload, organization=org)
        serializer.is_valid(raise_exception=True)
        return serializer.validated_data, serializer.get_resolved_device()

    @patch('invoices.pipeline.submit_to_zatca')
    def test_invoice_type_code_388_sets_document_type_invoice(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)

        submission = process_invoice_submission(org, resolved_device, validated_data)

        self.assertEqual(submission.document_type, InvoiceSubmission.DOCUMENT_TYPE_INVOICE)
        self.assertEqual(InvoiceSubmission.objects.count(), 1)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_invoice_type_code_381_sets_document_type_credit_note(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        payload = {
            **VALID_PAYLOAD, 'invoice_type_code': '381', 'billing_reference': 'INV-001',
            'reason': 'Goods returned',
        }
        validated_data, resolved_device = self._validated(payload, org)

        submission = process_invoice_submission(org, resolved_device, validated_data)

        self.assertEqual(submission.document_type, InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE)
        self.assertEqual(InvoiceSubmission.objects.count(), 1)
        self.assertIn('<cbc:InstructionNote>Goods returned</cbc:InstructionNote>', submission.xml_document)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_invoice_type_code_383_sets_document_type_debit_note(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        payload = {
            **VALID_PAYLOAD, 'invoice_type_code': '383', 'billing_reference': 'INV-001',
            'reason': 'Additional charges',
        }
        validated_data, resolved_device = self._validated(payload, org)

        submission = process_invoice_submission(org, resolved_device, validated_data)

        self.assertEqual(submission.document_type, InvoiceSubmission.DOCUMENT_TYPE_DEBIT_NOTE)
        self.assertEqual(InvoiceSubmission.objects.count(), 1)


@override_settings(DEVICE_KEY_ENCRYPTION_KEY=Fernet.generate_key().decode())
class ReturnInvoiceFlowTests(TestCase):
    """Exercises the return-invoice (credit note) flow, both via services.py directly and via the API."""

    def _make_org_with_signing_device(self, **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        org = Organization.objects.create(**defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device

    def _validated(self, payload, org):
        serializer = InvoiceSubmissionSerializer(data=payload, organization=org)
        serializer.is_valid(raise_exception=True)
        return serializer.validated_data, serializer.get_resolved_device()

    def _auth_header(self, org):
        return {'HTTP_AUTHORIZATION': f'ApiKey {org.api_key}'}

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_without_system_return_number_auto_generates_cn_number(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)
        self.assertEqual(invoice.icv, 1)

        credit_note = create_return_credit_note(org, device, invoice, reason='damaged goods')

        self.assertEqual(credit_note.icv, 2)
        self.assertEqual(credit_note.document_type, InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE)
        self.assertEqual(credit_note.original_invoice_id, invoice.pk)
        self.assertEqual(credit_note.system_return_number, '')
        self.assertEqual(credit_note.invoice_number, f'CN-{credit_note.icv}')
        self.assertEqual(credit_note.payload['invoice_number'], f'CN-{credit_note.icv}')
        self.assertEqual(credit_note.payload['billing_reference'], invoice.payload['invoice_number'])
        self.assertEqual(credit_note.status, InvoiceSubmission.STATUS_SUBMITTED)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_with_system_return_number_uses_it_as_invoice_number(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        credit_note = create_return_credit_note(
            org, device, invoice, system_return_number='SYS-99', reason='damaged goods',
        )

        self.assertEqual(credit_note.system_return_number, 'SYS-99')
        self.assertEqual(credit_note.invoice_number, 'SYS-99')
        self.assertEqual(credit_note.payload['invoice_number'], 'SYS-99')
        self.assertEqual(credit_note.payload['billing_reference'], invoice.payload['invoice_number'])
        self.assertEqual(credit_note.status, InvoiceSubmission.STATUS_SUBMITTED)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_with_duplicate_system_return_number_raises(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)
        create_return_credit_note(org, device, invoice, system_return_number='SYS-1')

        with self.assertRaises(DuplicateReturnNumberError):
            create_return_credit_note(org, device, invoice, system_return_number='SYS-1')

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_via_api_creates_credit_note(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        response = self.client.post(
            f'/api/invoices/{invoice.pk}/return/',
            data=json.dumps({'system_return_number': 'SYS-1', 'reason': 'test'}),
            content_type='application/json',
            **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            InvoiceSubmission.objects.filter(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE).count(), 1,
        )
        credit_note = InvoiceSubmission.objects.get(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE)
        self.assertEqual(credit_note.original_invoice_id, invoice.pk)
        self.assertEqual(credit_note.system_return_number, 'SYS-1')
        self.assertEqual(credit_note.invoice_number, 'SYS-1')

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_via_api_with_duplicate_system_return_number_returns_400(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)
        create_return_credit_note(org, device, invoice, system_return_number='SYS-1')

        response = self.client.post(
            f'/api/invoices/{invoice.pk}/return/',
            data=json.dumps({'system_return_number': 'SYS-1'}),
            content_type='application/json',
            **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('system_return_number', response.json())

    @patch('invoices.pipeline.submit_to_zatca')
    def test_return_via_api_cross_org_returns_404(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(VALID_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        other_org = Organization.objects.create(
            name='Other Org', branch_name='B', industry_category='IT',
            vat_number='399999999900044', country_code='SA',
            national_address_code='X', street_name='Y', building_number='2',
            city_sub_division='D', city_name='Jeddah', postal_zone='11111',
            cr_number='9999999998', invoice_category='1100', is_active=True,
        )

        response = self.client.post(
            f'/api/invoices/{invoice.pk}/return/',
            data=json.dumps({}),
            content_type='application/json',
            **self._auth_header(other_org),
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(
            InvoiceSubmission.objects.filter(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE).count(), 0,
        )


@override_settings(DEVICE_KEY_ENCRYPTION_KEY=Fernet.generate_key().decode())
class CustomReturnInvoiceFlowTests(TestCase):
    """Exercises the custom (partial) return flow: excluding items, editing
    qty/price, and defaulting/overriding the issue date."""

    TWO_ITEM_PAYLOAD = {
        **VALID_PAYLOAD,
        'items': [
            {'slno': 1, 'code': 'ITEM-001', 'name': 'Consultation', 'qty': '1.0000',
             'price': '100.0000', 'vat_type': 'S'},
            {'slno': 2, 'code': 'ITEM-002', 'name': 'Medicine', 'qty': '2.0000',
             'price': '50.0000', 'vat_type': 'S'},
        ],
    }

    def _make_org_with_signing_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device, user

    def _validated(self, payload, org):
        serializer = InvoiceSubmissionSerializer(data=payload, organization=org)
        serializer.is_valid(raise_exception=True)
        return serializer.validated_data, serializer.get_resolved_device()

    def _formset_post_data(self, include_second=False):
        data = {
            'form-TOTAL_FORMS': '2',
            'form-INITIAL_FORMS': '2',
            'form-MIN_NUM_FORMS': '0',
            'form-MAX_NUM_FORMS': '1000',
            'form-0-slno': '1', 'form-0-code': 'ITEM-001', 'form-0-name': 'Consultation',
            'form-0-vat_type': 'S', 'form-0-include': 'on', 'form-0-qty': '1.0000', 'form-0-price': '100.0000',
            'form-1-slno': '2', 'form-1-code': 'ITEM-002', 'form-1-name': 'Medicine',
            'form-1-vat_type': 'S', 'form-1-qty': '2.0000', 'form-1-price': '50.0000',
            'issue_date': '2026-07-20',
            'system_return_number': '',
            'reason': 'partial return',
        }
        if include_second:
            data['form-1-include'] = 'on'
        return data

    @patch('invoices.pipeline.submit_to_zatca')
    def test_excluding_an_item_leaves_it_out_of_the_credit_note(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device, _user = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        credit_note = create_custom_return_credit_note(
            org, device, invoice,
            items=[invoice.payload['items'][0]],
            issue_date=date(2026, 7, 20),
            reason='partial return',
        )

        self.assertEqual(len(credit_note.payload['items']), 1)
        self.assertEqual(credit_note.payload['items'][0]['code'], 'ITEM-001')
        self.assertEqual(credit_note.payload['issue_date'], '2026-07-20')
        self.assertNotEqual(credit_note.payload['issue_date'], invoice.payload['issue_date'])
        self.assertEqual(credit_note.document_type, InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE)
        self.assertEqual(credit_note.original_invoice_id, invoice.pk)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_edited_qty_and_price_flow_through_to_the_credit_note(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device, _user = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        edited_item = {**invoice.payload['items'][0], 'qty': '1.0000', 'price': '40.0000'}
        credit_note = create_custom_return_credit_note(
            org, device, invoice, items=[edited_item], issue_date=date(2026, 7, 20), reason='partial refund',
        )

        self.assertEqual(credit_note.payload['items'][0]['price'], '40.0000')
        self.assertNotEqual(credit_note.payload['items'][0]['price'], invoice.payload['items'][0]['price'])

    @patch('invoices.pipeline.submit_to_zatca')
    def test_duplicate_system_return_number_raises(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device, _user = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)
        create_custom_return_credit_note(
            org, device, invoice, items=[invoice.payload['items'][0]], issue_date=date(2026, 7, 20),
            system_return_number='SYS-CUSTOM-1',
        )

        with self.assertRaises(DuplicateReturnNumberError):
            create_custom_return_credit_note(
                org, device, invoice, items=[invoice.payload['items'][0]], issue_date=date(2026, 7, 20),
                system_return_number='SYS-CUSTOM-1',
            )

    @patch('invoices.pipeline.submit_to_zatca')
    def test_view_get_defaults_issue_date_to_today(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device, user = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        self.client.force_login(user)
        response = self.client.get(reverse('organization:invoice-return-custom', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['form'].initial['issue_date'], timezone.localdate())

    @patch('invoices.pipeline.submit_to_zatca')
    def test_view_post_creates_partial_credit_note_with_custom_date(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device, user = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        self.client.force_login(user)
        response = self.client.post(
            reverse('organization:invoice-return-custom', args=[org.pk, invoice.pk]),
            self._formset_post_data(include_second=False),
        )

        self.assertRedirects(response, reverse('organization:invoice-list', args=[org.pk]))
        credit_note = InvoiceSubmission.objects.get(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE)
        self.assertEqual(len(credit_note.payload['items']), 1)
        self.assertEqual(credit_note.payload['items'][0]['code'], 'ITEM-001')
        self.assertEqual(credit_note.payload['issue_date'], '2026-07-20')

    @patch('invoices.pipeline.submit_to_zatca')
    def test_view_post_with_no_items_selected_shows_error_and_creates_nothing(self, mock_submit):
        mock_submit.return_value = {'status_code': 200}
        org, device, user = self._make_org_with_signing_device()
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        self.client.force_login(user)
        post_data = self._formset_post_data(include_second=False)
        del post_data['form-0-include']
        response = self.client.post(
            reverse('organization:invoice-return-custom', args=[org.pk, invoice.pk]), post_data,
        )

        self.assertEqual(response.status_code, 422)
        self.assertEqual(
            InvoiceSubmission.objects.filter(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE).count(), 0,
        )

    @patch('invoices.pipeline.submit_to_zatca')
    def test_view_zatca_rejection_shows_message_and_creates_no_row(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        mock_submit.return_value = {'status_code': 200}
        validated_data, resolved_device = self._validated(self.TWO_ITEM_PAYLOAD, org)
        invoice = process_invoice_submission(org, resolved_device, validated_data)

        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        self.client.force_login(user)
        response = self.client.post(
            reverse('organization:invoice-return-custom', args=[org.pk, invoice.pk]),
            self._formset_post_data(include_second=False),
        )

        self.assertRedirects(response, reverse('organization:invoice-list', args=[org.pk]))
        self.assertEqual(
            InvoiceSubmission.objects.filter(document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE).count(), 0,
        )


class InvoiceNumbersByDateViewTests(TestCase):

    def _make_org_with_device(self, **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        org = Organization.objects.create(**defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        return org, device

    def _auth_header(self, org):
        return {'HTTP_AUTHORIZATION': f'ApiKey {org.api_key}'}

    def _make_submission(self, org, device, document_type, invoice_number, issue_date, icv):
        return InvoiceSubmission.objects.create(
            organization=org,
            device=device,
            document_type=document_type,
            invoice_number=invoice_number,
            payload={'invoice_number': invoice_number, 'issue_date': issue_date},
            status=InvoiceSubmission.STATUS_SUBMITTED,
            icv=icv,
        )

    def test_returns_all_document_types_when_no_filter_given(self):
        org, device = self._make_org_with_device()
        self._make_submission(org, device, 'invoice', 'INV-001', '2026-06-24', 1)
        self._make_submission(org, device, 'credit_note', 'CN-2', '2026-06-24', 2)
        self._make_submission(org, device, 'debit_note', 'DN-3', '2026-06-24', 3)
        self._make_submission(org, device, 'invoice', 'INV-OTHER-DAY', '2026-06-23', 4)

        response = self.client.get(
            '/api/invoices/numbers/', {'date': '2026-06-24'}, **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body['document_type'], 'all')
        self.assertEqual(body['count'], 3)
        self.assertEqual(set(body['invoice_numbers']), {'INV-001', 'CN-2', 'DN-3'})

    def test_filters_by_document_type(self):
        org, device = self._make_org_with_device()
        self._make_submission(org, device, 'invoice', 'INV-001', '2026-06-24', 1)
        self._make_submission(org, device, 'credit_note', 'CN-2', '2026-06-24', 2)

        response = self.client.get(
            '/api/invoices/numbers/', {'date': '2026-06-24', 'document_type': 'credit_note'},
            **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body['document_type'], 'credit_note')
        self.assertEqual(body['invoice_numbers'], ['CN-2'])

    def test_returns_empty_list_for_date_with_no_submissions(self):
        org, device = self._make_org_with_device()
        self._make_submission(org, device, 'invoice', 'INV-001', '2026-06-24', 1)

        response = self.client.get(
            '/api/invoices/numbers/', {'date': '2026-01-01'}, **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body['count'], 0)
        self.assertEqual(body['invoice_numbers'], [])

    def test_missing_date_returns_400(self):
        org, _ = self._make_org_with_device()

        response = self.client.get('/api/invoices/numbers/', **self._auth_header(org))

        self.assertEqual(response.status_code, 400)
        self.assertIn('date', response.json())

    def test_invalid_document_type_returns_400(self):
        org, _ = self._make_org_with_device()

        response = self.client.get(
            '/api/invoices/numbers/', {'date': '2026-06-24', 'document_type': 'bogus'},
            **self._auth_header(org),
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('document_type', response.json())

    def test_does_not_leak_other_organizations_invoices(self):
        org, device = self._make_org_with_device()
        other_org, other_device = self._make_org_with_device(
            vat_number='399999999900055', cr_number='9999999955',
        )
        self._make_submission(org, device, 'invoice', 'INV-MINE', '2026-06-24', 1)
        self._make_submission(other_org, other_device, 'invoice', 'INV-OTHER', '2026-06-24', 1)

        response = self.client.get(
            '/api/invoices/numbers/', {'date': '2026-06-24'}, **self._auth_header(org),
        )

        self.assertEqual(response.json()['invoice_numbers'], ['INV-MINE'])

    def test_missing_api_key_returns_401(self):
        response = self.client.get('/api/invoices/numbers/', {'date': '2026-06-24'})

        self.assertEqual(response.status_code, 401)


class InvoiceResubmitViewTests(TestCase):

    def _make_org_with_signing_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device, user

    def _validated(self, payload, org):
        serializer = InvoiceSubmissionSerializer(data=payload, organization=org)
        serializer.is_valid(raise_exception=True)
        return serializer.validated_data, serializer.get_resolved_device()

    def _make_not_submitted_invoice(self, org, device, icv=1):
        """Legacy fixture: a not_submitted row with an already-signed XML, as this
        app used to produce before submission became fully atomic (see CLAUDE.md's
        "Legacy not_submitted rows"). process_invoice_submission() no longer creates
        rows like this on rejection, so tests of the legacy Resubmit path build one
        directly rather than relying on the pipeline to produce it."""
        return InvoiceSubmission.objects.create(
            organization=org,
            device=device,
            document_type=InvoiceSubmission.DOCUMENT_TYPE_INVOICE,
            invoice_number=VALID_PAYLOAD['invoice_number'],
            payload={**VALID_PAYLOAD},
            status=InvoiceSubmission.STATUS_NOT_SUBMITTED,
            icv=icv,
            invoice_uuid=uuid.uuid4(),
            xml_document='<Invoice>stub</Invoice>',
            invoice_hash='stub-hash',
        )

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_not_submitted_invoice_marks_submitted(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        invoice = self._make_not_submitted_invoice(org, device)

        mock_submit.return_value = {'status_code': 200}
        self.client.force_login(user)
        response = self.client.post(reverse('organization:invoice-resubmit', args=[org.pk, invoice.pk]))

        self.assertRedirects(response, reverse('organization:invoice-list', args=[org.pk]))
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, InvoiceSubmission.STATUS_SUBMITTED)
        self.assertIsNotNone(invoice.submitted_at)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_keeps_not_submitted_on_repeated_failure(self, mock_submit):
        mock_submit.return_value = {'status_code': 400, 'error': 'boom'}
        org, device, user = self._make_org_with_signing_device()
        invoice = self._make_not_submitted_invoice(org, device)

        self.client.force_login(user)
        response = self.client.post(reverse('organization:invoice-resubmit', args=[org.pk, invoice.pk]))

        self.assertRedirects(response, reverse('organization:invoice-list', args=[org.pk]))
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, InvoiceSubmission.STATUS_NOT_SUBMITTED)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_already_submitted_invoice_does_not_repost(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        invoice = self._make_not_submitted_invoice(org, device)
        invoice.status = InvoiceSubmission.STATUS_SUBMITTED
        invoice.save(update_fields=['status'])

        self.client.force_login(user)
        response = self.client.post(reverse('organization:invoice-resubmit', args=[org.pk, invoice.pk]))

        self.assertRedirects(response, reverse('organization:invoice-list', args=[org.pk]))
        mock_submit.assert_not_called()

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_cross_owner_returns_404(self, mock_submit):
        org, device, _user = self._make_org_with_signing_device()
        invoice = self._make_not_submitted_invoice(org, device)

        _other_org, _other_device, other_user = self._make_org_with_signing_device(
            email='other@example.com', vat_number='399999999900099', cr_number='9999999997',
        )
        self.client.force_login(other_user)

        response = self.client.post(reverse('organization:invoice-resubmit', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 404)
        mock_submit.assert_not_called()

    def test_resubmit_anonymous_redirected_to_login(self):
        org, device, _user = self._make_org_with_signing_device()
        invoice = self._make_not_submitted_invoice(org, device)

        response = self.client.post(reverse('organization:invoice-resubmit', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)


class FailedSubmissionViewTests(TestCase):

    def _make_org_with_signing_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        private_key = ec_module.generate_private_key(ec_module.SECP256R1())
        pem = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        ).decode('ascii')
        DeviceKeyMaterial.objects.create(device=device, private_key_pem=encrypt_private_key(pem))
        return org, device, user

    def _make_failure(self, org, device, mock_submit, **payload_overrides):
        mock_submit.return_value = {'status_code': 422, 'error': {'message': 'invalid'}}
        payload = {**VALID_PAYLOAD, **payload_overrides}
        serializer = InvoiceSubmissionSerializer(data=payload, organization=org)
        serializer.is_valid(raise_exception=True)
        with self.assertRaises(InvoiceSubmissionRejected) as ctx:
            process_invoice_submission(org, serializer.get_resolved_device(), serializer.validated_data)
        return ctx.exception.failure

    @patch('invoices.pipeline.submit_to_zatca')
    def test_list_defaults_to_unresolved_only(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        unresolved = self._make_failure(org, device, mock_submit)
        resolved = self._make_failure(org, device, mock_submit, invoice_number='INV-002')
        resolved.resolved = True
        resolved.save(update_fields=['resolved'])
        self.client.force_login(user)

        response = self.client.get(reverse('organization:failed-submission-list', args=[org.pk]))

        self.assertEqual([f.pk for f in response.context['failures']], [unresolved.pk])

    @patch('invoices.pipeline.submit_to_zatca')
    def test_list_filters_by_invoice_number(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        self._make_failure(org, device, mock_submit, invoice_number='AAA-1')
        self._make_failure(org, device, mock_submit, invoice_number='BBB-1')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:failed-submission-list', args=[org.pk]), {'invoice_number': 'AAA'},
        )

        self.assertEqual([f.invoice_number for f in response.context['failures']], ['AAA-1'])

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_happy_path_marks_resolved(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        mock_submit.return_value = {'status_code': 200}
        self.client.force_login(user)

        response = self.client.post(
            reverse('organization:failed-submission-resubmit', args=[org.pk, failure.pk]),
        )

        self.assertRedirects(response, reverse('organization:failed-submission-list', args=[org.pk]))
        failure.refresh_from_db()
        self.assertTrue(failure.resolved)
        self.assertIsNotNone(failure.resolved_submission)
        self.assertEqual(failure.resolved_submission.status, InvoiceSubmission.STATUS_SUBMITTED)
        self.assertEqual(failure.resolved_submission.icv, 1)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_still_rejected_stays_unresolved_and_logs_new_failure(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        self.client.force_login(user)

        response = self.client.post(
            reverse('organization:failed-submission-resubmit', args=[org.pk, failure.pk]),
        )

        self.assertRedirects(response, reverse('organization:failed-submission-list', args=[org.pk]))
        failure.refresh_from_db()
        self.assertFalse(failure.resolved)
        self.assertEqual(InvoiceSubmissionFailure.objects.count(), 2)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_invalid_payload_shows_errors_without_submitting(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        failure.payload = {**failure.payload, 'items': []}
        failure.save(update_fields=['payload'])
        self.client.force_login(user)
        mock_submit.reset_mock()

        response = self.client.post(
            reverse('organization:failed-submission-resubmit', args=[org.pk, failure.pk]),
        )

        self.assertRedirects(response, reverse('organization:failed-submission-list', args=[org.pk]))
        failure.refresh_from_db()
        self.assertFalse(failure.resolved)
        mock_submit.assert_not_called()

    @patch('invoices.pipeline.submit_to_zatca')
    def test_resubmit_cross_owner_returns_404(self, mock_submit):
        org, device, _user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        _other_org, _other_device, other_user = self._make_org_with_signing_device(
            email='other@example.com', vat_number='399999999900098', cr_number='9999999996',
        )
        self.client.force_login(other_user)

        response = self.client.post(
            reverse('organization:failed-submission-resubmit', args=[org.pk, failure.pk]),
        )

        self.assertEqual(response.status_code, 404)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_delete_confirm_page_shows_failure(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:failed-submission-delete', args=[org.pk, failure.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['object'], failure)

    @patch('invoices.pipeline.submit_to_zatca')
    def test_delete_removes_unresolved_failure(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        self.client.force_login(user)

        response = self.client.post(reverse('organization:failed-submission-delete', args=[org.pk, failure.pk]))

        self.assertRedirects(response, reverse('organization:failed-submission-list', args=[org.pk]))
        self.assertFalse(InvoiceSubmissionFailure.objects.filter(pk=failure.pk).exists())

    @patch('invoices.pipeline.submit_to_zatca')
    def test_delete_removes_resolved_failure(self, mock_submit):
        org, device, user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        mock_submit.return_value = {'status_code': 200}
        self.client.force_login(user)
        self.client.post(reverse('organization:failed-submission-resubmit', args=[org.pk, failure.pk]))
        failure.refresh_from_db()
        self.assertTrue(failure.resolved)

        response = self.client.post(reverse('organization:failed-submission-delete', args=[org.pk, failure.pk]))

        self.assertRedirects(response, reverse('organization:failed-submission-list', args=[org.pk]))
        self.assertFalse(InvoiceSubmissionFailure.objects.filter(pk=failure.pk).exists())

    @patch('invoices.pipeline.submit_to_zatca')
    def test_delete_cross_owner_returns_404(self, mock_submit):
        org, device, _user = self._make_org_with_signing_device()
        failure = self._make_failure(org, device, mock_submit)
        _other_org, _other_device, other_user = self._make_org_with_signing_device(
            email='deleteother@example.com', vat_number='399999999900093', cr_number='9999999991',
        )
        self.client.force_login(other_user)

        response = self.client.post(reverse('organization:failed-submission-delete', args=[org.pk, failure.pk]))

        self.assertEqual(response.status_code, 404)
        self.assertTrue(InvoiceSubmissionFailure.objects.filter(pk=failure.pk).exists())


class InvoiceListViewTests(TestCase):

    def _make_org_with_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        return org, device, user

    def _make_submission(self, org, device, icv, **payload_overrides):
        payload = {
            'invoice_number': f'INV-{icv:03d}',
            'issue_date': timezone.localdate().isoformat(),
            'customer_name': 'Test Customer',
        }
        payload.update(payload_overrides)
        return InvoiceSubmission.objects.create(
            organization=org,
            device=device,
            document_type=payload_overrides.get('document_type', InvoiceSubmission.DOCUMENT_TYPE_INVOICE),
            invoice_number=payload['invoice_number'],
            payload=payload,
            status=payload_overrides.get('status', InvoiceSubmission.STATUS_SUBMITTED),
            icv=icv,
        )

    def test_filters_by_invoice_number(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, invoice_number='INV-AAA')
        self._make_submission(org, device, 2, invoice_number='INV-BBB')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'invoice_number': 'AAA'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([i.icv for i in response.context['invoices']], [1])

    def test_filters_by_customer_name(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, customer_name='Acme Corp')
        self._make_submission(org, device, 2, customer_name='Beta LLC')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'customer_name': 'acme'},
        )

        self.assertEqual([i.icv for i in response.context['invoices']], [1])

    def test_filters_by_issue_date_range(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, issue_date='2026-06-24')
        self._make_submission(org, device, 2, issue_date='2026-06-25')
        self._make_submission(org, device, 3, issue_date='2026-06-26')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]),
            {'issue_date_from': '2026-06-25', 'issue_date_to': '2026-06-25'},
        )

        self.assertEqual([i.icv for i in response.context['invoices']], [2])

    def test_filters_by_issue_date_open_ended_range(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, issue_date='2026-06-24')
        self._make_submission(org, device, 2, issue_date='2026-06-25')
        self._make_submission(org, device, 3, issue_date='2026-06-26')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'issue_date_from': '2026-06-25'},
        )

        self.assertEqual({i.icv for i in response.context['invoices']}, {2, 3})

    def test_defaults_to_todays_invoices_on_cold_navigation(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1)
        self._make_submission(org, device, 2, issue_date='2026-06-24')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-list', args=[org.pk]))

        self.assertEqual([i.icv for i in response.context['invoices']], [1])
        today = timezone.localdate().isoformat()
        self.assertEqual(response.context['filters']['issue_date_from'], today)
        self.assertEqual(response.context['filters']['issue_date_to'], today)

    def test_filters_by_icv(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1)
        self._make_submission(org, device, 2)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-list', args=[org.pk]), {'icv': '2'})

        self.assertEqual([i.icv for i in response.context['invoices']], [2])

    def test_filters_by_document_type(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, document_type=InvoiceSubmission.DOCUMENT_TYPE_INVOICE)
        self._make_submission(org, device, 2, document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE)
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'document_type': 'credit_note'},
        )

        self.assertEqual([i.icv for i in response.context['invoices']], [2])

    def test_filters_by_status(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, status=InvoiceSubmission.STATUS_SUBMITTED)
        self._make_submission(org, device, 2, status=InvoiceSubmission.STATUS_NOT_SUBMITTED)
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'status': 'not_submitted'},
        )

        self.assertEqual([i.icv for i in response.context['invoices']], [2])

    def test_pagination_splits_across_pages(self):
        org, device, user = self._make_org_with_device()
        for icv in range(1, 31):
            self._make_submission(org, device, icv)
        self.client.force_login(user)

        page1 = self.client.get(reverse('organization:invoice-list', args=[org.pk]))
        page2 = self.client.get(reverse('organization:invoice-list', args=[org.pk]), {'page': 2})

        self.assertEqual(len(page1.context['invoices']), 25)
        self.assertTrue(page1.context['is_paginated'])
        self.assertEqual(len(page2.context['invoices']), 5)

    def test_does_not_leak_other_organizations_invoices(self):
        org, device, user = self._make_org_with_device()
        other_org, other_device, _other_user = self._make_org_with_device(
            email='other@example.com', vat_number='399999999900098', cr_number='9999999996',
        )
        self._make_submission(org, device, 1)
        self._make_submission(other_org, other_device, 1)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-list', args=[org.pk]))

        self.assertEqual(len(response.context['invoices']), 1)

    def _make_submission_with_items(self, org, device, icv, **payload_overrides):
        payload_overrides.setdefault('items', [
            {'slno': 1, 'code': 'ITEM-001', 'name': 'Widget', 'qty': '1', 'price': '100', 'vat_type': 'S'},
        ])
        return self._make_submission(org, device, icv, **payload_overrides)

    def test_summary_defaults_to_page_scope(self):
        org, device, user = self._make_org_with_device()
        for icv in range(1, 31):
            self._make_submission_with_items(org, device, icv)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-list', args=[org.pk]))

        self.assertEqual(response.context['summary_scope'], 'page')
        self.assertEqual(response.context['summary']['count'], 25)
        self.assertEqual(response.context['summary']['total_amount'], Decimal('2500.00'))

    def test_summary_all_scope_sums_every_filtered_row(self):
        org, device, user = self._make_org_with_device()
        for icv in range(1, 31):
            self._make_submission_with_items(org, device, icv)
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'summary_scope': 'all'},
        )

        self.assertEqual(response.context['summary_scope'], 'all')
        self.assertEqual(response.context['summary']['count'], 30)
        self.assertEqual(response.context['summary']['total_amount'], Decimal('3000.00'))

    def test_summary_scope_invalid_falls_back_to_page(self):
        org, device, user = self._make_org_with_device()
        for icv in range(1, 31):
            self._make_submission_with_items(org, device, icv)
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-list', args=[org.pk]), {'summary_scope': 'bogus'},
        )

        self.assertEqual(response.context['summary_scope'], 'page')
        self.assertEqual(response.context['summary']['count'], 25)

    def test_summary_nets_credit_notes_and_adds_debit_notes(self):
        org, device, user = self._make_org_with_device()
        self._make_submission_with_items(org, device, 1)
        self._make_submission_with_items(org, device, 2)
        self._make_submission_with_items(
            org, device, 3, document_type=InvoiceSubmission.DOCUMENT_TYPE_CREDIT_NOTE,
        )
        self._make_submission_with_items(
            org, device, 4, document_type=InvoiceSubmission.DOCUMENT_TYPE_DEBIT_NOTE,
            items=[{'slno': 1, 'code': 'ITEM-001', 'name': 'Widget', 'qty': '1', 'price': '50', 'vat_type': 'S'}],
        )
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-list', args=[org.pk]))

        # count stays a plain row count (unnetted); the monetary total nets
        # the two invoices (100 each) plus the debit note (50) minus the
        # credit note (100): 100 + 100 + 50 - 100 = 150.
        self.assertEqual(response.context['summary']['count'], 4)
        self.assertEqual(response.context['summary']['total_amount'], Decimal('150.00'))

    def test_export_returns_filtered_rows_and_summary(self):
        org, device, user = self._make_org_with_device()
        self._make_submission_with_items(org, device, 1, issue_date='2026-06-24')
        self._make_submission_with_items(org, device, 2, issue_date='2026-06-25')
        self._make_submission_with_items(org, device, 3, issue_date='2026-06-25')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-export', args=[org.pk]),
            {'issue_date_from': '2026-06-25', 'issue_date_to': '2026-06-25'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename=', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0][0], 'Type')
        # 2 matching invoices (icv 2 and 3) + header + summary row
        self.assertEqual(len(rows), 4)
        invoice_numbers = {rows[1][1], rows[2][1]}
        self.assertEqual(invoice_numbers, {'INV-002', 'INV-003'})
        self.assertEqual(rows[3][0], 'Summary (2 invoices)')
        self.assertEqual(rows[3][4], Decimal('200.00'))

    def test_export_includes_rows_beyond_a_single_page(self):
        org, device, user = self._make_org_with_device()
        for icv in range(1, 31):
            self._make_submission_with_items(org, device, icv)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-export', args=[org.pk]))

        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        # header + 30 invoices + summary row
        self.assertEqual(len(rows), 32)

    def test_export_does_not_leak_other_organizations_invoices(self):
        org, device, user = self._make_org_with_device()
        other_org, other_device, _other_user = self._make_org_with_device(
            email='exportother@example.com', vat_number='399999999900097', cr_number='9999999995',
        )
        self._make_submission_with_items(org, device, 1)
        self._make_submission_with_items(other_org, other_device, 1)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-export', args=[org.pk]))

        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(len(rows), 3)  # header + 1 invoice + summary


class InvoiceXmlZipExportTests(TestCase):

    def _make_org_with_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        return org, device, user

    def _make_submission(self, org, device, icv, xml_document='', issue_date=None, **payload_overrides):
        issue_date = issue_date or timezone.localdate().isoformat()
        payload = {'invoice_number': f'INV-{icv:03d}', 'issue_date': issue_date, 'customer_name': 'Test Customer'}
        payload.update(payload_overrides)
        return InvoiceSubmission.objects.create(
            organization=org,
            device=device,
            document_type=payload_overrides.get('document_type', InvoiceSubmission.DOCUMENT_TYPE_INVOICE),
            invoice_number=payload['invoice_number'],
            payload=payload,
            status=InvoiceSubmission.STATUS_SUBMITTED,
            icv=icv,
            xml_document=xml_document,
        )

    def test_zip_contains_only_filtered_rows(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, xml_document='<A/>', issue_date='2026-06-24')
        self._make_submission(org, device, 2, xml_document='<B/>', issue_date='2026-06-25')
        self._make_submission(org, device, 3, xml_document='<C/>', issue_date='2026-06-25')
        self.client.force_login(user)

        response = self.client.get(
            reverse('organization:invoice-xml-zip-export', args=[org.pk]),
            {'issue_date_from': '2026-06-25', 'issue_date_to': '2026-06-25'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/zip')
        self.assertIn('attachment; filename=', response['Content-Disposition'])

        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            names = zf.namelist()
            self.assertEqual(len(names), 2)
            contents = {zf.read(name).decode() for name in names}
            self.assertEqual(contents, {'<B/>', '<C/>'})

    def test_zip_skips_rows_with_no_xml_document(self):
        org, device, user = self._make_org_with_device()
        self._make_submission(org, device, 1, xml_document='<A/>')
        self._make_submission(org, device, 2, xml_document='')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml-zip-export', args=[org.pk]))

        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertEqual(len(zf.namelist()), 1)

    def test_zip_does_not_leak_other_organizations_invoices(self):
        org, device, user = self._make_org_with_device()
        other_org, other_device, _other_user = self._make_org_with_device(
            email='xmlzipother@example.com', vat_number='399999999900094', cr_number='9999999992',
        )
        self._make_submission(org, device, 1, xml_document='<A/>')
        self._make_submission(other_org, other_device, 1, xml_document='<Other/>')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml-zip-export', args=[org.pk]))

        with zipfile.ZipFile(BytesIO(response.content)) as zf:
            self.assertEqual(len(zf.namelist()), 1)


class InvoiceDetailViewTests(TestCase):

    def _make_org_with_device(self, email='owner@example.com', **org_overrides):
        defaults = {**ORG_DEFAULTS}
        defaults.update(org_overrides)
        user = User.objects.create_user(username=email, email=email, password='testpass123')
        org = Organization.objects.create(email=email, owner_user=user, **defaults)
        device = Device.objects.create(
            organization=org,
            asset_id='ASSET-100',
            egs_sw_serial_number='SERIAL-200',
            otp='123456',
            csid_response=FAKE_CSID,
        )
        return org, device, user

    def _make_submission(self, org, device, icv=1, qr_code_data='', xml_document='', **payload_overrides):
        payload = {
            'invoice_number': f'INV-{icv:03d}',
            'issue_date': '2026-07-14',
            'issue_time': '10:00:00',
            'invoice_type_code_name_attribute': '020000000',
            'customer_name': 'Test Customer',
            'customer_vat': '300000000000003',
            'customer_city': 'Riyadh',
            'customer_country_code': 'SA',
            'items': [
                {'slno': 1, 'code': 'ITEM-1', 'name': 'Widget', 'qty': '2', 'price': '50', 'vat_type': 'S'},
            ],
        }
        payload.update(payload_overrides)
        return InvoiceSubmission.objects.create(
            organization=org,
            device=device,
            document_type=payload_overrides.get('document_type', InvoiceSubmission.DOCUMENT_TYPE_INVOICE),
            invoice_number=payload['invoice_number'],
            payload=payload,
            status=InvoiceSubmission.STATUS_SUBMITTED,
            icv=icv,
            qr_code_data=qr_code_data,
            xml_document=xml_document,
        )

    def test_detail_shows_line_items_and_totals(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-detail', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 200)
        line_items = response.context['line_items']
        self.assertEqual(len(line_items), 1)
        self.assertEqual(line_items[0]['line_amount'], Decimal('100.00'))
        self.assertEqual(line_items[0]['line_vat'], Decimal('15.00'))
        self.assertEqual(line_items[0]['line_total'], Decimal('115.00'))
        self.assertEqual(response.context['invoice'].net_with_tax, Decimal('115.00'))

    def test_detail_renders_qr_image_when_qr_code_data_present(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device, qr_code_data='AQhUZXN0IENv')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-detail', args=[org.pk, invoice.pk]))

        self.assertTrue(response.context['qr_image'].startswith('data:image/png;base64,'))

    def test_detail_qr_image_none_when_not_yet_available(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device, qr_code_data='')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-detail', args=[org.pk, invoice.pk]))

        self.assertIsNone(response.context['qr_image'])

    def test_detail_does_not_leak_other_organizations_invoice(self):
        org, device, user = self._make_org_with_device()
        other_org, other_device, _other_user = self._make_org_with_device(
            email='detailother@example.com', vat_number='399999999900096', cr_number='9999999994',
        )
        other_invoice = self._make_submission(other_org, other_device)
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-detail', args=[org.pk, other_invoice.pk]))

        self.assertEqual(response.status_code, 404)

    def test_detail_anonymous_redirected_to_login(self):
        org, device, _user = self._make_org_with_device()
        invoice = self._make_submission(org, device)

        response = self.client.get(reverse('organization:invoice-detail', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_view_xml_serves_inline(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device, xml_document='<Invoice>content</Invoice>')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/xml')
        self.assertIn('inline; filename=', response['Content-Disposition'])
        self.assertEqual(response.content.decode(), '<Invoice>content</Invoice>')

    def test_download_xml_serves_as_attachment(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device, xml_document='<Invoice>content</Invoice>')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml-download', args=[org.pk, invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertIn('attachment; filename=', response['Content-Disposition'])
        self.assertEqual(response.content.decode(), '<Invoice>content</Invoice>')

    def test_view_xml_redirects_with_message_when_xml_not_available(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device, xml_document='')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml', args=[org.pk, invoice.pk]))

        self.assertRedirects(response, reverse('organization:invoice-detail', args=[org.pk, invoice.pk]))

    def test_download_xml_redirects_with_message_when_xml_not_available(self):
        org, device, user = self._make_org_with_device()
        invoice = self._make_submission(org, device, xml_document='')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml-download', args=[org.pk, invoice.pk]))

        self.assertRedirects(response, reverse('organization:invoice-detail', args=[org.pk, invoice.pk]))

    def test_view_xml_does_not_leak_other_organizations_invoice(self):
        org, device, user = self._make_org_with_device()
        other_org, other_device, _other_user = self._make_org_with_device(
            email='xmlother@example.com', vat_number='399999999900095', cr_number='9999999993',
        )
        other_invoice = self._make_submission(other_org, other_device, xml_document='<Invoice/>')
        self.client.force_login(user)

        response = self.client.get(reverse('organization:invoice-xml', args=[org.pk, other_invoice.pk]))

        self.assertEqual(response.status_code, 404)
